import functools
import pprint
import random
import re
import string
import time
from pathlib import Path
from itertools import count

import requests
from bs4 import BeautifulSoup as bs
from fake_useragent import UserAgent
from openpyxl import Workbook

class Writer:
	def __init__(self, filename):
		self.filename = filename

class XlsxWriter(Writer):
	def __init__(self, fields, filename='output'):
		super().__init__(filename)
		self.fields = fields 
		self.letters = string.ascii_uppercase[:len(self.fields)]
		self.file_type = '.xlsx'
		self.check_filename()
		self.open_an_active_sheet()
		self.write_sheet_headers()

	def __repr__(self):
		return self.filename

	def check_filename(self):
		if self.file_type not in self.filename:
			self.filename += self.file_type
	
	def open_an_active_sheet(self):
		self.workbook = Workbook()
		self.sheet = self.workbook.active

	def close_workbook(self):
		self.workbook.save(filename=self.filename)

	def write_sheet_headers(self):
		for letter, field in zip(self.letters, self.fields):
			self.sheet[letter + str(self.sheet.max_row)].value = field

	def write_to_sheet(self, dictionary):
		try:
			max_row = str(self.sheet.max_row + 1)
			for letter, field in zip(self.letters, self.fields):
				self.sheet[letter + max_row].value = dictionary.get(field).encode('utf-8', errors='ignore')
		finally:
			self.close_workbook()

class FileReader:	
	@property
	def content(self):
		path_object = Path(input("\aEnter a valid filename: "))
		if path_object.exists():
			with path_object.open() as file_handler:
				return [ line.strip() for line in file_handler.readlines() if line not in string.whitespace ]
		raise Exception("\aYou might have to check the file name.")

class NtnSnrLocators:
	number_of_results = '.BottomHeaderSlot h1'
	product_list_items = 'app-ntn-product-list-item'
	product_list_item = 'li'
	name_brand = 'div[class*="name-brand"]'

def yield_count():
	for _ in count:
		yield str(_)

def sleep(secs=random.randint(1, 3)):
	time.sleep(secs)

def make_header():
	return {
		'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
		'Accept-Encoding': 'gzip, deflate, br',
		'Accept-Language': 'en-US,en;q=0.9',
		'Cache-Control': 'max-age=0',
		'Connection': 'keep-alive',
		'Host': 'eshop.ntn-snr.com',
		'Upgrade-Insecure-Requests': '1',
		'User-Agent': ua.random,
	}

def retry(function):
	'''tries to run a function after an unsuccessful attempt.'''
	@functools.wraps(function)
	def inner(*args, **kwargs):
		for _ in range(5):
			try:
				return function(*args, **kwargs)	
			except Exception as err:
				print(err)
	return inner

@retry
def make_request(string):
	# make a list out of the string and a regex of the first numbers in 
	# the string, use this to run a search
	for number in {string, re.search('\d*', string).group()}:
		print(f'Searching {number}')
		url = 'https://eshop.ntn-snr.com/en/search/' + number  + '?tabNum=1&searchQueryContext=COMPETITOR_DEFAULT&matchType=CONTAINS'
		response = requests.get(url, headers=make_header(), timeout=60)
		if response.ok:
			# proceed to making a soup if response is ok
			soup = bs(response.text, features='html.parser')
			# check if there's result
			if (product_list_items := soup.select(NtnSnrLocators.product_list_items)):
				return product_list_items
		else:
			print(response.reason)

def prepopulate_dict(default='N/A'):
	'''
	make dictionary with default values of 'N/A'
	'''
	default_values = [default] * len(fields)
	return dict(list( zip(fields, default_values) ))

def sift_data(items, number):
	if items:
		for item in items:
			data = prepopulate_dict('-')
			data['Input'] = number
			data['Sr. No'] = next(yield_count())
			
			lis = item.select(NtnSnrLocators.product_list_item)
			for li in lis:
				split = li.text.split()
				# print(split)
				data[split[1].strip().upper()] = split[0].strip()

			# name brand
			split = item.select_one(NtnSnrLocators.name_brand).text.split('-')
			data[split[1].strip().upper()] = split[0].strip()
			# write data to file
			pprint.pprint(data)
			writer.write_to_sheet(data)
	else:
		print(text)

def main():
	for number in FileReader().content:
		items = make_request(number)
		if items:
			sift_data(items, number)
		sleep()  # some seconds delay between each search 
	input('Press ENter to close app...')

if __name__ == '__main__':
	fields = [
		'Sr. No',
		'Input',
		'SNR',
		'NTN',
		'SKF',
		'FAG',
		'NSK',
		'TIMKEN'
	]
	writer = XlsxWriter(fields)
	count = count(1)
	ua = UserAgent(fallback='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36')
	main()
	writer.close_workbook()