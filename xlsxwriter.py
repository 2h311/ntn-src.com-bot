import string

from openpyxl import Workbook

class Writer:
	def __init__(self, filename):
		self.filename = filename

class XlsxWriter(Writer):
	# TODO: enable a context manager
	def __init__(self, fields, filename='output'):
		super().__init__(filename)
		self.fields = fields 
		self.letters = string.ascii_uppercase[:len(self.fields)]
		self.file_type = '.xlsx'
		self.check_filename()
		self.open_an_active_sheet()
		self.write_sheet_headers()

	def __repr__(self):
		return self.filename

	def check_filename(self):
		if self.file_type not in self.filename:
			self.filename += self.file_type
	
	def open_an_active_sheet(self):
		self.workbook = Workbook()
		self.sheet = self.workbook.active

	def close_workbook(self):
		self.workbook.save(filename=self.filename)

	def write_sheet_headers(self):
		for letter, field in zip(self.letters, self.fields):
			self.sheet[letter + str(self.sheet.max_row)].value = field

	def write_to_sheet(self, dictionary):
		try:
			max_row = str(self.sheet.max_row + 1)
			for letter, field in zip(self.letters, self.fields):
				self.sheet[letter + max_row].value = dictionary.get(field).encode('utf-8', errors='ignore')
		finally:
			self.close_workbook()
